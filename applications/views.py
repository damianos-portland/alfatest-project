from django.shortcuts import render

from .forms import *
from .models import *
from .filters import *


# Extra Imports for the Login and Logout Capabilities
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.core.paginator import Paginator
import xlwt
import os
import xlwings
import shutil
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook




from django.contrib.auth.models import User
#from django.contrib.auth.decorators import login_required



def all_customers(request):

    customers = Customer.objects.all()

    #orders = customer.order_set.all()
    total_customers = customers.count()


    mycustFilter = CustomerFilter(request.GET, queryset=customers)
    customers=mycustFilter.qs

    has_filter = any(field in request.GET for field in set(mycustFilter.get_fields()))


    paginator = Paginator(customers, 5) # Show 25 contacts per page.

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    #tableFilter = OrderFilter(request.GET, queryset=orders)
    #orders = myFilter.qs
    context = {'has_filter':has_filter,'page_obj': page_obj, 'customers': customers, 'mycustFilter':mycustFilter}
    return render(request, 'applications/all_customers.html', context)



def newcustomer(request):

    registered = False

    if request.method == 'POST':

        # Get info from "both" forms
        # It appears as one form to the user on the .html page

        customer_form = CustomerForm(data=request.POST)
        #profile_form = UserProfileInfoForm(data=request.POST)

        # Check to see both forms are valid
        if customer_form.is_valid() :

            # Save User Form to Database

            customer = customer_form.save()
            registered = True

        else:
            # One of the forms was invalid if this else gets called.
            print('One of the forms was invalid if this else gets called.')

    else:
        # Was not an HTTP post so we just render the forms as blank.

        customer_form = CustomerForm()


    # This is the render and context dictionary to feed
    # back to the registration.html file page.
    return render(request,'applications/newcustomer.html',
                          {'customer_form':customer_form,
                           'registered':registered})
# def newapp(request) :
#     return render(request,'applications/newapp.html')


def new_application(request):

    registered = False

    if request.method == 'POST':

        # Get info from "both" forms
        # It appears as one form to the user on the .html page
        application_form = ApplicationForm(data=request.POST)
        customer_form = CustomerForm(data=request.POST)
        #profile_form = UserProfileInfoForm(data=request.POST)

        print(application_form.data)
        print(request.POST)


        # Check to see both forms are valid
        if (application_form.is_valid()) :

            # Save User Form to Database

            application = application_form.save()
            registered = True

        else:
            # One of the forms was invalid if this else gets called.
            print('One of the forms was invalid if this else gets called.')

    else:
        # Was not an HTTP post so we just render the forms as blank.
        application_form = ApplicationForm()
        customer_form = CustomerForm()


    # This is the render and context dictionary to feed
    # back to the registration.html file page.
    return render(request,'applications/newapp.html',
                          {'application_form':application_form,'customer_form':customer_form,
                           'registered':registered})


def all_applications(request):

    applications = Application.objects.all()

    #orders = customer.order_set.all()
    total_applications = applications.count()
    delivered = applications.filter(status='Delivered').count()
    pending = applications.filter(status='Pending').count()

    myFilter = OrderFilter(request.GET, queryset=applications)
    applications=myFilter.qs

    has_filter = any(field in request.GET for field in set(myFilter.get_fields()))


    paginator = Paginator(applications, 5) # Show 25 contacts per page.

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    #tableFilter = OrderFilter(request.GET, queryset=orders)
    #orders = myFilter.qs
    context = {'has_filter':has_filter,'page_obj': page_obj, 'applications': applications, 'delivered': delivered, 'pendind': pending , 'myFilter':myFilter}
    return render(request, 'applications/all_applications.html', context)



def updateOrder(request, pk):

    registered = False
    order = Application.objects.get(id=pk)
    form = UpdateApplicationForm(instance=order)


    if request.method == 'POST':
        form = UpdateApplicationForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            registered = True

    context = {'form':form, 'registered':registered}
    return render(request, 'applications/order_form.html', context)



def updateCustomer(request, pk):

    registered = False
    order = Customer.objects.get(id=pk)
    form = CustomerForm(instance=order)


    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            registered = True

    context = {'form':form, 'registered':registered}
    return render(request, 'applications/customer_form.html', context)


def createXL(request,pk):

    order = Application.objects.get(id=pk)
    # response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="finalapplication.xls"'
    #
    #
    #
    #
    # shutil.copy("C:\\Users\\User\\Desktop\\ALFATEST\\Files\\application.xlsx", "C:\\Users\\User\\Desktop\\ALFATEST\\Files\\newFile.xlsx")
    #
    # wb = xlwings.Book("C:\\Users\\User\\Desktop\\ALFATEST\\Files\\newFile.xlsx")
    # Sheet1 = wb.sheets[0]
    # # Sheet2 = wb.sheets[7]
    # #Then update as you want it
    # Sheet1.range(9, 12).value = order.customer.name #This will change the cell(2,4) to 4
    # wb.save()
    shutil.copy("C:\\Users\\User\\Desktop\\ALFATEST\\Files\\application.xlsx", "C:\\Users\\User\\Desktop\\ALFATEST\\Files\\newFile.xlsx")
    wb = load_workbook("C:\\Users\\User\\Desktop\\ALFATEST\\Files\\newFile.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]

    #Then update as you want it
    Sheet1 .cell(row = 9, column = 12).value = order.customer.name
    Sheet1 .cell(row = 2, column = 28).value = order.protocol_number
    Sheet1 .cell(row = 10, column = 12).value = order.customer.address
    Sheet1 .cell(row = 12, column = 12).value = order.customer.afm
    Sheet1 .cell(row = 13, column = 12).value = order.customer.phone
    Sheet1 .cell(row = 14, column = 12).value = order.customer.engineer
    # Sheet1 .cell(row = 15, column = 12).value = order.date
    Sheet1 .cell(row = 4, column = 28).value = order.date
    Sheet1 .cell(row = 12, column = 28).value = order.customer.doy
    Sheet1 .cell(row = 13, column = 28).value = order.customer.fax

    counter=0

    for k in order.trials.all() :
        Sheet1 .cell(row = 28+counter, column = 2).value = k.name
        counter+=1








     #This will change the cell(2,4) to 4
    # wb.save("C:\\Users\\User\\Desktop\\ALFATEST\\Files\\finalnewFile.xlsx")

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=Application_{order.protocol_number}.xlsx'
    return response







def exportXL(request, pk):

    # registered = False
    order = Application.objects.get(id=pk)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="application.xlsx"'
    # filename="C:\\Users\\User\\Desktop\\users1.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

        # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['protocol_number', 'ergo', 'customer', 'date', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    row = [order.protocol_number,order.ergo.eidos,order.customer.name,order.date]
    row_num += 1
    for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response



def load_trials(request):
    yliko_id = request.GET.get('ergo')
    trials = Trial.objects.filter(yliko_id=yliko_id).order_by('name')
    return render(request, 'applications/dropdown_list_trials.html', {'trials': trials, 'test':yliko_id})
